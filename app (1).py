# app.py
from flask import Flask, request, redirect, render_template_string
import sqlite3
from typing import List, Tuple, Optional, Dict
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import os

app = Flask(__name__)

# -----------------------------
# 0) PATH & FILE CONFIG
# -----------------------------
BASE_DIR = Path(__file__).resolve().parent
DB_FILE = str(BASE_DIR / "tasks.db")
EXCEL_FILE = str(BASE_DIR / "task_done_log.xlsx")

# -----------------------------
# 1) DATABASE HELPERS
# -----------------------------
def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            person TEXT,
            category TEXT,
            done_count INTEGER DEFAULT 0
        )""")

def ensure_columns():
    with sqlite3.connect(DB_FILE) as conn:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(tasks)")]
        if "category" not in cols:
            conn.execute("ALTER TABLE tasks ADD COLUMN category TEXT")
        if "done_count" not in cols:
            conn.execute("ALTER TABLE tasks ADD COLUMN done_count INTEGER DEFAULT 0")

def add_task(name: str, person: Optional[str], category: Optional[str]):
    name = (name or "").strip()[:120]
    person = (person or "").strip()[:60] or None
    category = (category or "").strip().upper() or "UNCATEGORIZED"
    if not name:
        return
    with sqlite3.connect(DB_FILE) as conn:
        cur = conn.execute("INSERT INTO tasks (name, person, category) VALUES (?, ?, ?)",
                           (name, person, category))
        tid = cur.lastrowid
    log_to_excel("create", tid, name, person, None, category)

def delete_tasks(ids: List[str]):
    if not ids:
        return
    qmarks = ",".join("?" * len(ids))
    with sqlite3.connect(DB_FILE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(f"SELECT * FROM tasks WHERE id IN ({qmarks})", ids).fetchall()
        conn.executemany("DELETE FROM tasks WHERE id = ?", [(x,) for x in ids])
    for r in rows:
        log_to_excel("delete", r["id"], r["name"], r["person"], None, r["category"])

def mark_tasks_done(ids: List[str]):
    if not ids:
        return
    qmarks = ",".join("?" * len(ids))
    with sqlite3.connect(DB_FILE) as conn:
        conn.row_factory = sqlite3.Row
        for tid in ids:
            conn.execute("UPDATE tasks SET done_count = COALESCE(done_count,0)+1 WHERE id=?", (tid,))
        rows = conn.execute(f"SELECT * FROM tasks WHERE id IN ({qmarks})", ids).fetchall()
        conn.executemany("DELETE FROM tasks WHERE id=?", [(x,) for x in ids])
    for r in rows:
        log_to_excel("done", r["id"], r["name"], r["person"], r["done_count"], r["category"])

def fetch_tasks() -> List[Tuple[int, str, Optional[str], Optional[str]]]:
    with sqlite3.connect(DB_FILE) as conn:
        return conn.execute("SELECT id,name,person,category FROM tasks ORDER BY id DESC").fetchall()

# -----------------------------
# 2) EXCEL LOGGING
# -----------------------------
def log_to_excel(action, tid, name, person, done_count=None, category=None):
    cat = (category or "UNCATEGORIZED").upper()
    path = Path(EXCEL_FILE)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
    # ชีตรวม
    if "Tasks_Log" not in wb.sheetnames:
        ws = wb.create_sheet("Tasks_Log", 0)
        ws.append(["timestamp", "action", "task_id", "name", "person", "done_count", "category"])
    else:
        ws = wb["Tasks_Log"]
    now = datetime.now().isoformat(timespec="seconds")
    ws.append([now, action, tid or "", name, person or "", done_count or "", cat])
    # แยกชีตตามหมวด
    valid = {"MALE OPERATION", "PROJECT"}
    sheet = cat if cat in valid else "UNCATEGORIZED"
    if sheet not in wb.sheetnames:
        ws2 = wb.create_sheet(sheet)
        ws2.append(["timestamp", "action", "task_id", "name", "person", "done_count"])
    else:
        ws2 = wb[sheet]
    ws2.append([now, action, tid or "", name, person or "", done_count or ""])
    wb.save(path)

# -----------------------------
# 3) GROUP TASKS BY PERSON
# -----------------------------
def group_tasks_by_person(tasks):
    groups = defaultdict(list)
    for t in tasks:
        tid, name, person, cat = t
        groups[person or "ไม่ระบุ"].append(t)
    result = []
    for i, (person, items) in enumerate(sorted(groups.items())):
        result.append({"gid": i, "person": person, "items": items})
    return result

# -----------------------------
# 4) HTML TEMPLATE
# -----------------------------
HTML = """
<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<title>งานวันนี้</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  :root { --brand:#d9480f; --ink:#222; --muted:#666; }
  body { font-family: system-ui, -apple-system, 'Segoe UI', Tahoma, sans-serif;
         background:#000; color:var(--ink);
         max-width:1200px; margin:48px auto; padding:24px; }

  h1 { text-align:center; color:var(--brand); margin-bottom:20px; }

  .card { background:#fff; padding:16px; border-radius:12px;
          box-shadow:0 1px 6px rgba(0,0,0,.08); margin-bottom:16px; }

  label { display:block; font-size:14px; margin-bottom:6px; }

  /* ✅ แยกฟอนต์แต่ละกล่อง input */
  .task-input {
    width:80%; padding:9px 10px; font-size:18px;
    border:1.5px solid #ccc; border-radius:10px; margin-bottom:12px;
  }
  .task-input:focus { border-color:var(--brand); }

  .person-input {
    width:100%; padding:8px 10px; font-size:15px;
    border:1.5px solid #ccc; border-radius:10px;
  }
  .person-input:focus { border-color:var(--brand); }

  .category-select {
    width:100%; padding:8px 10px; font-size:14px;
    border:1.5px solid #ccc; border-radius:10px;
  }
  .category-select:focus { border-color:var(--brand); }

  button {
    width:100%; padding:12px; background:var(--brand); color:#fff;
    font-size:16px; border:none; border-radius:10px; cursor:pointer;
  }
  button:hover { filter:brightness(.95); }

  .row { display:grid; grid-template-columns:1fr 1fr; gap:10px; }

  .task { display:flex; justify-content:space-between; align-items:center;
          background:#fff; padding:10px 12px; border:1px solid #eee;
          border-radius:10px; margin:8px 0; }
  .info { display:flex; flex-direction:column; }
  .person { font-size:13px; color:#555; }

  .person-group-container {
    display:flex; flex-wrap:wrap; gap:16px; justify-content:center;
  }
  .person-group-container .card {
    flex:1 1 280px; min-width:260px;
  }
</style>
</head>
<body>
  <h1>งานวันนี้</h1>

  <!-- เพิ่มงาน -->
  <div class="card">
    <form method="POST">
      <input type="hidden" name="action" value="add">
      <label>ชื่องาน</label>
      <input type="text" name="new_task" class="task-input" required placeholder="กรอกชื่องาน">
      <label>ผู้รับผิดชอบ</label>
      <div class="row">
        <input type="text" name="person" class="person-input" placeholder="ชื่อผู้รับผิดชอบ">
        <select name="category" class="category-select" required>
          <option value="MALE OPERATION">MALE OPERATION</option>
          <option value="PROJECT">PROJECT</option>
        </select>
      </div>
      <button type="submit">เพิ่มงาน</button>
    </form>
  </div>

  <!-- รายการงาน -->
  <form method="POST" id="listForm">
    <input type="hidden" name="action" value="done">
    {% if groups|length == 0 %}
      <div class="card"><div style="text-align:center;color:#666;">ยังไม่มีงาน</div></div>
    {% else %}
      <div class="person-group-container">
        {% for g in groups %}
        <div class="card">
          <div><b>{{ g["person"] }}</b></div>
          {% for t in g["items"] %}
          {% set tid=t[0] %}{% set name=t[1] %}{% set cat=t[3] %}
          <div class="task">
            <div class="info">
              <div><b>{{ loop.index }}.</b> {{ name }}</div>
              <div class="person">หมวด: {{ cat }}</div>
            </div>
            <input type="checkbox" name="ids" value="{{ tid }}">
          </div>
          {% endfor %}
        </div>
        {% endfor %}
      </div>
      <div style="display:flex;gap:10px;margin-top:10px;">
        <button class="secondary" type="submit">บันทึกงานที่เสร็จ</button>
        <button class="danger" formaction="/delete" formmethod="POST">ลบงาน</button>
      </div>
    {% endif %}
  </form>
</body>
</html>
"""

# -----------------------------
# 5) ROUTES
# -----------------------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        action = (request.form.get("action") or "").lower()
        if action == "add":
            add_task(request.form.get("new_task"), request.form.get("person"), request.form.get("category"))
        elif action == "done":
            mark_tasks_done(request.form.getlist("ids"))
        return redirect("/")
    tasks = fetch_tasks()
    groups = group_tasks_by_person(tasks)
    return render_template_string(HTML, groups=groups)

@app.route("/delete", methods=["POST"])
def delete_route():
    delete_tasks(request.form.getlist("ids"))
    return redirect("/")

# -----------------------------
# 6) BOOTSTRAP
# -----------------------------
def bootstrap():
    os.makedirs(BASE_DIR, exist_ok=True)
    init_db()
    ensure_columns()

bootstrap()

# -----------------------------
# 7) ENTRY POINT
# -----------------------------
if __name__ == "__main__":
    print("✅ Server running at http://127.0.0.1:5050")
    app.run(debug=True, host="127.0.0.1", port=5050)
